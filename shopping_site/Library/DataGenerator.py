import openpyxl


def generate_data():
    wb = openpyxl.load_workbook(".\\Assets\\email&password.xlsx")
    sh = wb["Sheet"]
    rows = sh.max_row
    columns = sh.max_column
    print(rows)
    print(columns)
    lis = []
    for i in range(1, rows):
        lis2 = []
        for j in range(1, columns + 1):
            lis2.append(sh.cell(i, j).value)
        lis.append(lis2)
    return lis
